import ipaddress, io, csv
from typing import Optional, List
from datetime import datetime

from fastapi import FastAPI, Depends, HTTPException, Query, UploadFile, File
from fastapi.staticfiles import StaticFiles
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from sqlalchemy.orm import Session
import openpyxl

from database import get_db, init_db, Subnet

app = FastAPI(title="IPAM", version="1.0.0")

app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

@app.on_event("startup")
def startup():
    init_db()

# ── Schemas ──────────────────────────────────────────────────────────────────
class SubnetIn(BaseModel):
    site: str
    subnet: str
    vlan: Optional[str] = ""
    description: Optional[str] = ""
    site_type: Optional[str] = "LAN"
    status: Optional[str] = "active"
    notes: Optional[str] = ""

# ── Helpers ───────────────────────────────────────────────────────────────────
def cidr_host_count(cidr: str) -> int:
    try:
        return ipaddress.ip_network(cidr, strict=False).num_addresses
    except Exception:
        return 0

def validate_cidr(cidr: str):
    try:
        ipaddress.ip_network(cidr, strict=False)
    except ValueError:
        raise HTTPException(status_code=422, detail=f"Invalid CIDR notation: {cidr}")

def enrich(row: Subnet) -> dict:
    d = {c.name: getattr(row, c.name) for c in row.__table__.columns}
    d["host_count"] = cidr_host_count(row.subnet)
    return d

# ── Subnet CRUD ───────────────────────────────────────────────────────────────
@app.get("/api/subnets", response_model=List[dict])
def list_subnets(
    site: Optional[str] = None,
    site_type: Optional[str] = None,
    status: Optional[str] = None,
    vlan: Optional[str] = None,
    q: Optional[str] = None,
    skip: int = 0,
    limit: int = 2000,
    db: Session = Depends(get_db)
):
    qs = db.query(Subnet)
    if site:       qs = qs.filter(Subnet.site == site)
    if site_type:  qs = qs.filter(Subnet.site_type == site_type)
    if status:     qs = qs.filter(Subnet.status == status)
    if vlan:       qs = qs.filter(Subnet.vlan == vlan)
    if q:
        like = f"%{q}%"
        qs = qs.filter(
            Subnet.site.ilike(like) |
            Subnet.subnet.ilike(like) |
            Subnet.description.ilike(like) |
            Subnet.notes.ilike(like) |
            Subnet.vlan.ilike(like)
        )
    rows = qs.order_by(Subnet.site, Subnet.id).offset(skip).limit(limit).all()
    return [enrich(r) for r in rows]

@app.get("/api/subnets/{subnet_id}")
def get_subnet(subnet_id: int, db: Session = Depends(get_db)):
    row = db.query(Subnet).filter(Subnet.id == subnet_id).first()
    if not row:
        raise HTTPException(404, "Not found")
    return enrich(row)

@app.post("/api/subnets", status_code=201)
def create_subnet(data: SubnetIn, db: Session = Depends(get_db)):
    validate_cidr(data.subnet)
    existing = db.query(Subnet).filter(Subnet.subnet == data.subnet, Subnet.site == data.site).first()
    if existing:
        raise HTTPException(409, f"Subnet {data.subnet} already exists for site {data.site}")
    row = Subnet(**data.dict())
    db.add(row)
    db.commit()
    db.refresh(row)
    return enrich(row)

@app.put("/api/subnets/{subnet_id}")
def update_subnet(subnet_id: int, data: SubnetIn, db: Session = Depends(get_db)):
    row = db.query(Subnet).filter(Subnet.id == subnet_id).first()
    if not row:
        raise HTTPException(404, "Not found")
    validate_cidr(data.subnet)
    conflict = db.query(Subnet).filter(
        Subnet.subnet == data.subnet,
        Subnet.site == data.site,
        Subnet.id != subnet_id
    ).first()
    if conflict:
        raise HTTPException(409, f"Subnet {data.subnet} already exists for site {data.site}")
    for k, v in data.dict().items():
        setattr(row, k, v)
    row.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(row)
    return enrich(row)

@app.delete("/api/subnets/{subnet_id}", status_code=204)
def delete_subnet(subnet_id: int, db: Session = Depends(get_db)):
    row = db.query(Subnet).filter(Subnet.id == subnet_id).first()
    if not row:
        raise HTTPException(404, "Not found")
    db.delete(row)
    db.commit()

# ── Stats ─────────────────────────────────────────────────────────────────────
@app.get("/api/stats")
def stats(db: Session = Depends(get_db)):
    all_rows = db.query(Subnet).all()
    total = len(all_rows)
    by_type   = {}
    by_site   = {}
    by_status = {}
    for r in all_rows:
        by_type[r.site_type]  = by_type.get(r.site_type, 0) + 1
        by_site[r.site]       = by_site.get(r.site, 0) + 1
        by_status[r.status]   = by_status.get(r.status, 0) + 1
    return {
        "total": total,
        "by_type": by_type,
        "by_site": by_site,
        "by_status": by_status,
        "sites": sorted(set(r.site for r in all_rows)),
        "types": sorted(set(r.site_type for r in all_rows)),
        "vlans": sorted(set(r.vlan for r in all_rows if r.vlan))
    }

# ── Export CSV ────────────────────────────────────────────────────────────────
@app.get("/api/export/csv")
def export_csv(db: Session = Depends(get_db)):
    rows = db.query(Subnet).order_by(Subnet.site, Subnet.id).all()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID","Site","Subnet","VLAN","Description","Type","Status","Notes","Created","Updated"])
    for r in rows:
        writer.writerow([r.id, r.site, r.subnet, r.vlan, r.description,
                         r.site_type, r.status, r.notes,
                         r.created_at.strftime("%Y-%m-%d %H:%M"),
                         r.updated_at.strftime("%Y-%m-%d %H:%M")])
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=ipam_export.csv"}
    )

# ── Export Excel ──────────────────────────────────────────────────────────────
@app.get("/api/export/xlsx")
def export_xlsx(db: Session = Depends(get_db)):
    rows = db.query(Subnet).order_by(Subnet.site, Subnet.id).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IPAM"
    ws.append(["ID","Site","Subnet","VLAN","Description","Type","Status","Notes","Created","Updated"])
    for r in rows:
        ws.append([r.id, r.site, r.subnet, r.vlan or "", r.description or "",
                   r.site_type, r.status, r.notes or "",
                   r.created_at.strftime("%Y-%m-%d %H:%M"),
                   r.updated_at.strftime("%Y-%m-%d %H:%M")])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=ipam_export.xlsx"}
    )

# ── Import CSV/XLSX ───────────────────────────────────────────────────────────
@app.post("/api/import")
async def import_file(file: UploadFile = File(...), db: Session = Depends(get_db)):
    content = await file.read()
    added = 0
    skipped = 0
    errors = []

    def process_row(row_dict):
        nonlocal added, skipped
        site   = str(row_dict.get("Site","")).strip()
        subnet = str(row_dict.get("Subnet","")).strip()
        if not site or not subnet:
            return
        try:
            ipaddress.ip_network(subnet, strict=False)
        except ValueError:
            errors.append(f"Invalid CIDR: {subnet}")
            return
        if db.query(Subnet).filter(Subnet.subnet == subnet, Subnet.site == site).first():
            skipped += 1
            return
        db.add(Subnet(
            site=site, subnet=subnet,
            vlan=str(row_dict.get("VLAN","")).strip(),
            description=str(row_dict.get("Description","")).strip(),
            site_type=str(row_dict.get("Type","LAN")).strip() or "LAN",
            status=str(row_dict.get("Status","active")).strip() or "active",
            notes=str(row_dict.get("Notes","")).strip()
        ))
        added += 1

    fname = file.filename.lower()
    if fname.endswith(".csv"):
        reader = csv.DictReader(io.StringIO(content.decode("utf-8-sig")))
        for row in reader:
            process_row(row)
    elif fname.endswith(".xlsx"):
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        hdrs = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            process_row(dict(zip(hdrs, row)))
    else:
        raise HTTPException(400, "Only CSV or XLSX files supported")

    db.commit()
    return {"added": added, "skipped": skipped, "errors": errors}

# ── Serve frontend ─────────────────────────────────────────────────────────────
import os

def _resolve_frontend_dir():
    here = os.path.dirname(os.path.abspath(__file__))
    for candidate in [
        os.path.join(here, "frontend"),          # deploy layout: /opt/ipam/frontend/
        os.path.join(here, "..", "frontend"),     # dev layout:    backend/../frontend/
    ]:
        if os.path.isdir(candidate):
            return os.path.abspath(candidate)
    raise RuntimeError(f"Cannot find frontend directory near {here}")

FRONTEND_DIR = _resolve_frontend_dir()
STATIC_DIR   = os.path.join(FRONTEND_DIR, "static")
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")

@app.get("/", include_in_schema=False)
@app.get("/{full_path:path}", include_in_schema=False)
def serve_frontend(full_path: str = ""):
    return FileResponse(os.path.join(FRONTEND_DIR, "index.html"))
